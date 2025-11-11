from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from .models import Categoria, Subcategoria
from .forms import CategoriaForm, SubcategoriaForm

def is_superuser(user):
    return user.is_superuser

# Vista para Categorías
@login_required
@user_passes_test(is_superuser)
def categoria_list(request):
    """Lista todas las categorías"""
    categorias = Categoria.objects.all()
    return render(request, 'categoria/categoria_list.html', {'categorias': categorias})

@login_required
@user_passes_test(is_superuser)
def categoria_create(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('categoria:list')
    else:
        form = CategoriaForm()
    return render(request, 'categoria/categoria_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def categoria_edit(request, pk):
    """Editar categoría existente"""
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('categoria:list')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'categoria/categoria_form.html', {'form': form, 'categoria': categoria, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def categoria_delete(request, pk):
    """Eliminar categoría"""
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.')
        return redirect('categoria:list')
    return render(request, 'categoria/categoria_confirm_delete.html', {'categoria': categoria})

# Vista para Subcategorías
@login_required
@user_passes_test(is_superuser)
def subcategoria_list(request):
    """Lista todas las subcategorías"""
    subcategorias = Subcategoria.objects.all()
    return render(request, 'categoria/subcategoria_list.html', {'subcategorias': subcategorias})

@login_required
@user_passes_test(is_superuser)
def subcategoria_create(request):
    """Crear nueva subcategoría"""
    if request.method == 'POST':
        form = SubcategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subcategoría creada exitosamente.')
            return redirect('categoria:subcategoria_list')
    else:
        form = SubcategoriaForm()
    return render(request, 'categoria/subcategoria_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def subcategoria_edit(request, pk):
    """Editar subcategoría existente"""
    subcategoria = get_object_or_404(Subcategoria, pk=pk)
    if request.method == 'POST':
        form = SubcategoriaForm(request.POST, instance=subcategoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subcategoría actualizada exitosamente.')
            return redirect('categoria:subcategoria_list')
    else:
        form = SubcategoriaForm(instance=subcategoria)
    return render(request, 'categoria/subcategoria_form.html', {'form': form, 'subcategoria': subcategoria, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def subcategoria_delete(request, pk):
    """Eliminar subcategoría"""
    subcategoria = get_object_or_404(Subcategoria, pk=pk)
    if request.method == 'POST':
        subcategoria.delete()
        messages.success(request, 'Subcategoría eliminada exitosamente.')
        return redirect('categoria:subcategoria_list')
    return render(request, 'categoria/subcategoria_confirm_delete.html', {'subcategoria': subcategoria})

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def categoria_upload_excel(request):
    """Subir categorías desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                nombre = str(row[0]).strip() if row[0] else None
                if not nombre:
                    continue
                
                activo = True
                if len(row) > 1 and row[1] is not None:
                    activo_val = str(row[1]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    categoria, created = Categoria.objects.get_or_create(
                        nombre=nombre,
                        defaults={'activo': activo}
                    )
                    
                    if not created:
                        categoria.activo = activo
                        categoria.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            df = pd.read_excel(excel_file)
            
            for idx, row in df.iterrows():
                nombre = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if not nombre:
                    continue
                
                activo = True
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    activo_val = str(row.iloc[1]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    categoria, created = Categoria.objects.get_or_create(
                        nombre=nombre,
                        defaults={'activo': activo}
                    )
                    
                    if not created:
                        categoria.activo = activo
                        categoria.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def categoria_download_template(request):
    """Descargar template de Excel para categorías"""
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Categorías"
            
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            sheet['A1'] = 'Nombre'
            sheet['B1'] = 'Activo'
            
            for cell in ['A1', 'B1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            ejemplo_datos = [
                ['Hardware', 1],
                ['Software', 1],
                ['Redes', 1],
                ['Soporte Técnico', 0],
            ]
            
            for idx, (nombre, activo) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = nombre
                sheet[f'B{idx}'] = activo
            
            sheet.column_dimensions['A'].width = 35
            sheet.column_dimensions['B'].width = 12
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_categorias.xlsx"'
            workbook.save(response)
            return response
        else:
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Nombre': ['Hardware', 'Software', 'Redes', 'Soporte Técnico'],
                'Activo': [1, 1, 1, 0]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Categorías')
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_categorias.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def subcategoria_upload_excel(request):
    """Subir subcategorías desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                categoria_nombre = str(row[0]).strip() if row[0] else None
                categoria = None
                if categoria_nombre:
                    try:
                        categoria = Categoria.objects.get(nombre=categoria_nombre)
                    except Categoria.DoesNotExist:
                        errors.append(f"Fila {row_idx}: Categoría '{categoria_nombre}' no encontrada")
                        continue
                    except Categoria.MultipleObjectsReturned:
                        categoria = Categoria.objects.filter(nombre=categoria_nombre).first()
                
                nombre = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not nombre:
                    errors.append(f"Fila {row_idx}: Nombre es requerido")
                    continue
                
                activo = True
                if len(row) > 2 and row[2] is not None:
                    activo_val = str(row[2]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    subcategoria, created = Subcategoria.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'categoria': categoria,
                            'activo': activo
                        }
                    )
                    
                    if not created:
                        subcategoria.categoria = categoria
                        subcategoria.activo = activo
                        subcategoria.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            df = pd.read_excel(excel_file)
            
            for idx, row in df.iterrows():
                categoria_nombre = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                categoria = None
                if categoria_nombre:
                    try:
                        categoria = Categoria.objects.get(nombre=categoria_nombre)
                    except Categoria.DoesNotExist:
                        errors.append(f"Fila {idx + 2}: Categoría '{categoria_nombre}' no encontrada")
                        continue
                    except Categoria.MultipleObjectsReturned:
                        categoria = Categoria.objects.filter(nombre=categoria_nombre).first()
                
                nombre = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
                if not nombre:
                    errors.append(f"Fila {idx + 2}: Nombre es requerido")
                    continue
                
                activo = True
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    activo_val = str(row.iloc[2]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    subcategoria, created = Subcategoria.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'categoria': categoria,
                            'activo': activo
                        }
                    )
                    
                    if not created:
                        subcategoria.categoria = categoria
                        subcategoria.activo = activo
                        subcategoria.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def subcategoria_download_template(request):
    """Descargar template de Excel para subcategorías"""
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Subcategorías"
            
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            sheet['A1'] = 'Categoría'
            sheet['B1'] = 'Nombre'
            sheet['C1'] = 'Activo'
            
            for cell in ['A1', 'B1', 'C1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            ejemplo_datos = [
                ['Hardware', 'Monitores', 1],
                ['Hardware', 'Teclados', 1],
                ['Software', 'Sistemas Operativos', 1],
                ['Software', 'Aplicaciones', 1],
            ]
            
            for idx, (categoria, nombre, activo) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = categoria
                sheet[f'B{idx}'] = nombre
                sheet[f'C{idx}'] = activo
            
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 35
            sheet.column_dimensions['C'].width = 12
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_subcategorias.xlsx"'
            workbook.save(response)
            return response
        else:
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Categoría': ['Hardware', 'Hardware', 'Software', 'Software'],
                'Nombre': ['Monitores', 'Teclados', 'Sistemas Operativos', 'Aplicaciones'],
                'Activo': [1, 1, 1, 1]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Subcategorías')
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_subcategorias.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )
