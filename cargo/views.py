from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Cargo
from .forms import CargoForm

def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def cargo_list(request):
    """Lista todos los cargos"""
    cargos = Cargo.objects.all()
    return render(request, 'cargo/cargo_list.html', {'cargos': cargos})

@login_required
@user_passes_test(is_superuser)
def cargo_create(request):
    """Crear nuevo cargo"""
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo creado exitosamente.')
            return redirect('cargo:list')
    else:
        form = CargoForm()
    return render(request, 'cargo/cargo_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def cargo_edit(request, pk):
    """Editar cargo existente"""
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo actualizado exitosamente.')
            return redirect('cargo:list')
    else:
        form = CargoForm(instance=cargo)
    return render(request, 'cargo/cargo_form.html', {'form': form, 'cargo': cargo, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def cargo_delete(request, pk):
    """Eliminar cargo"""
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        cargo.delete()
        messages.success(request, 'Cargo eliminado exitosamente.')
        return redirect('cargo:list')
    return render(request, 'cargo/cargo_confirm_delete.html', {'cargo': cargo})

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def cargo_upload_excel(request):
    """Subir cargos desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        # Validar extensión
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        # Intentar importar openpyxl o pandas
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            # Usar openpyxl
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # Leer desde la fila 2 (asumiendo que la fila 1 tiene encabezados)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Si la primera columna está vacía, saltar
                    continue
                
                nombre = str(row[0]).strip() if row[0] else None
                if not nombre:
                    continue
                
                # Intentar obtener el estado (columna 2, si existe)
                activo = True
                if len(row) > 1 and row[1] is not None:
                    activo_val = str(row[1]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    cargo, created = Cargo.objects.get_or_create(
                        nombre=nombre,
                        defaults={'activo': activo}
                    )
                    
                    if not created:
                        # Actualizar si ya existe
                        cargo.activo = activo
                        cargo.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            # Usar pandas
            df = pd.read_excel(excel_file)
            
            # Asumir que la primera columna es "nombre" y la segunda (si existe) es "activo"
            for idx, row in df.iterrows():
                nombre = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if not nombre:
                    continue
                
                activo = True
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    activo_val = str(row.iloc[1]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    cargo, created = Cargo.objects.get_or_create(
                        nombre=nombre,
                        defaults={'activo': activo}
                    )
                    
                    if not created:
                        cargo.activo = activo
                        cargo.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]  # Limitar a 10 errores
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def cargo_download_template(request):
    """Descargar template de Excel para cargos"""
    try:
        # Intentar importar openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                from django.http import HttpResponse
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            # Crear workbook con openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Cargos"
            
            # Estilos
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Encabezados
            sheet['A1'] = 'Nombre'
            sheet['B1'] = 'Activo'
            
            # Aplicar estilos a los encabezados
            for cell in ['A1', 'B1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            # Datos de ejemplo
            ejemplo_datos = [
                ['Gerente General', 1],
                ['Analista de Sistemas', 1],
                ['Desarrollador', 1],
                ['Administrador de Base de Datos', 0],
            ]
            
            for idx, (nombre, activo) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = nombre
                sheet[f'B{idx}'] = activo
            
            # Ajustar ancho de columnas
            sheet.column_dimensions['A'].width = 35
            sheet.column_dimensions['B'].width = 12
            
            # Crear respuesta
            from django.http import HttpResponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_cargos.xlsx"'
            workbook.save(response)
            return response
        else:
            # Usar pandas
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Nombre': ['Gerente General', 'Analista de Sistemas', 'Desarrollador', 'Administrador de Base de Datos'],
                'Activo': [1, 1, 1, 0]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Cargos')
            
            from django.http import HttpResponse
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_cargos.xlsx"'
            return response
            
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )