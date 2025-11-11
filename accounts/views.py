from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.hashers import make_password
from .models import User, TelegramGrupoSede
from .forms import UserForm, UserCreateForm, ProfileForm
from .telegram_utils import (
    get_api_base, get_last_chat_id, send_text, 
    get_chat_info, delete_webhook, get_updates
)

def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def user_list(request):
    """Lista todos los usuarios"""
    users = User.objects.all()
    return render(request, 'accounts/user_list.html', {'users': users})

@login_required
@user_passes_test(is_superuser)
def user_create(request):
    """Crear nuevo usuario"""
    if request.method == 'POST':
        form = UserCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('accounts:list')
    else:
        form = UserCreateForm()
    return render(request, 'accounts/user_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def user_edit(request, pk):
    """Editar usuario existente"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente.')
            return redirect('accounts:list')
    else:
        form = UserForm(instance=user)
    return render(request, 'accounts/user_form.html', {'form': form, 'user': user, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def user_delete(request, pk):
    """Eliminar usuario"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Usuario eliminado exitosamente.')
        return redirect('accounts:list')
    return render(request, 'accounts/user_confirm_delete.html', {'user': user})

# Vistas para Telegram Bot
@login_required
def vincular_bot(request):
    """Vista para vincular el bot de Telegram al usuario actual"""
    user = request.user
    
    if request.method == 'POST':
        try:
            api_base = get_api_base()
            if not api_base:
                messages.error(request, 'Bot Token no configurado. Contacta al administrador.')
                return redirect('accounts:vincular_bot')
            
            # Obtener chat_id usando el DNI del usuario
            dni = user.dni
            if not dni:
                messages.error(request, 'Debes tener un DNI registrado para vincular el bot.')
                return redirect('accounts:vincular_bot')
            
            chat_id, updates = get_last_chat_id(api_base, dni=dni)
            
            if not chat_id:
                messages.warning(
                    request, 
                    f'No se encontró tu chat_id. Envía un mensaje al bot con tu DNI: {dni}'
                )
                return redirect('accounts:vincular_bot')
            
            # Verificar que el chat_id no esté usado por otro usuario
            if User.objects.filter(telegram_chat_id=chat_id).exclude(id=user.id).exists():
                messages.error(request, 'Este chat_id ya está vinculado a otro usuario.')
                return redirect('accounts:vincular_bot')
            
            # Guardar chat_id y vincular bot
            user.telegram_chat_id = str(chat_id)
            user.bot_vinculado = True
            user.save()
            
            # Enviar mensaje de confirmación
            try:
                send_text(api_base, chat_id, f'✅ Bot vinculado exitosamente para {user.nombre_completo}')
            except:
                pass  # Si falla el envío, no es crítico
            
            messages.success(request, 'Bot vinculado exitosamente.')
            return redirect('accounts:vincular_bot')
            
        except Exception as e:
            messages.error(request, f'Error al vincular bot: {str(e)}')
            return redirect('accounts:vincular_bot')
    
    # GET: Mostrar estado del bot
    context = {
        'bot_vinculado': user.bot_vinculado,
        'telegram_chat_id': user.telegram_chat_id,
        'dni': user.dni,
    }
    return render(request, 'accounts/vincular_bot.html', context)


@login_required
def desvincular_bot(request):
    """Vista para desvincular el bot de Telegram"""
    user = request.user
    
    if request.method == 'POST':
        user.bot_vinculado = False
        user.telegram_chat_id = None
        user.save()
        messages.success(request, 'Bot desvinculado exitosamente.')
        return redirect('accounts:vincular_bot')
    
    return redirect('accounts:vincular_bot')


@login_required
def obtener_chat_id(request):
    """API para obtener chat_id (AJAX)"""
    if request.method == 'POST':
        try:
            api_base = get_api_base()
            if not api_base:
                return JsonResponse({'error': 'Bot Token no configurado'}, status=400)
            
            dni = request.user.dni
            if not dni:
                return JsonResponse({'error': 'DNI no registrado'}, status=400)
            
            chat_id, updates = get_last_chat_id(api_base, dni=dni)
            
            if chat_id:
                return JsonResponse({
                    'success': True,
                    'chat_id': str(chat_id),
                    'message': 'Chat ID encontrado'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'No se encontró chat_id. Envía un mensaje al bot con tu DNI: {dni}'
                })
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# Vistas para gestión de grupos de Telegram por sede
@login_required
@user_passes_test(is_superuser)
def telegram_grupos_list(request):
    """Lista todos los grupos de Telegram por sede"""
    from sede.models import Sede
    
    grupos = TelegramGrupoSede.objects.select_related('sede').all()
    sedes_sin_grupo = Sede.objects.filter(activa=True).exclude(
        telegram_grupo__isnull=False
    )
    
    context = {
        'grupos': grupos,
        'sedes_sin_grupo': sedes_sin_grupo,
    }
    return render(request, 'accounts/telegram_grupos_list.html', context)


@login_required
@user_passes_test(is_superuser)
def telegram_grupo_create(request):
    """Crear grupo de Telegram para una sede"""
    from sede.models import Sede
    
    if request.method == 'POST':
        sede_id = request.POST.get('sede')
        grupo_chat_id = request.POST.get('grupo_chat_id')
        nombre_grupo = request.POST.get('nombre_grupo', '')
        
        if not sede_id or not grupo_chat_id:
            messages.error(request, 'Sede y Chat ID son requeridos.')
            return redirect('accounts:telegram_grupos_list')
        
        sede = get_object_or_404(Sede, id=sede_id)
        
        # Verificar que la sede no tenga ya un grupo
        if TelegramGrupoSede.objects.filter(sede=sede).exists():
            messages.error(request, f'La sede {sede.nombre} ya tiene un grupo asignado.')
            return redirect('accounts:telegram_grupos_list')
        
        # Crear grupo
        grupo = TelegramGrupoSede.objects.create(
            sede=sede,
            grupo_chat_id=grupo_chat_id,
            nombre_grupo=nombre_grupo
        )
        
        messages.success(request, f'Grupo creado exitosamente para {sede.nombre}.')
        return redirect('accounts:telegram_grupos_list')
    
    return redirect('accounts:telegram_grupos_list')


@login_required
@user_passes_test(is_superuser)
def telegram_grupo_edit(request, pk):
    """Editar grupo de Telegram"""
    from sede.models import Sede
    grupo = get_object_or_404(TelegramGrupoSede, pk=pk)
    
    if request.method == 'POST':
        sede_id = request.POST.get('sede')
        grupo_chat_id = request.POST.get('grupo_chat_id')
        nombre_grupo = request.POST.get('nombre_grupo', '')
        
        if not sede_id or not grupo_chat_id:
            messages.error(request, 'Sede y Chat ID son requeridos.')
            return redirect('accounts:telegram_grupos_list')
        
        sede = get_object_or_404(Sede, id=sede_id)
        
        # Verificar que la sede no tenga ya otro grupo (excepto el actual)
        otro_grupo = TelegramGrupoSede.objects.filter(sede=sede).exclude(pk=grupo.pk).first()
        if otro_grupo:
            messages.error(request, f'La sede {sede.nombre} ya tiene otro grupo asignado.')
            return redirect('accounts:telegram_grupos_list')
        
        # Actualizar grupo
        grupo.sede = sede
        grupo.grupo_chat_id = grupo_chat_id
        grupo.nombre_grupo = nombre_grupo
        grupo.save()
        
        messages.success(request, f'Grupo actualizado exitosamente para {sede.nombre}.')
        return redirect('accounts:telegram_grupos_list')
    
    # Obtener todas las sedes para el formulario
    sedes = Sede.objects.filter(activa=True).order_by('nombre')
    
    return render(request, 'accounts/telegram_grupo_edit.html', {
        'grupo': grupo,
        'sedes': sedes
    })


@login_required
@user_passes_test(is_superuser)
def telegram_grupo_delete(request, pk):
    """Eliminar grupo de Telegram"""
    grupo = get_object_or_404(TelegramGrupoSede, pk=pk)
    
    if request.method == 'POST':
        sede_nombre = grupo.sede.nombre
        grupo.delete()
        messages.success(request, f'Grupo eliminado exitosamente para {sede_nombre}.')
        return redirect('accounts:telegram_grupos_list')
    
    return render(request, 'accounts/telegram_grupo_confirm_delete.html', {'grupo': grupo})


@login_required
@user_passes_test(is_superuser)
def obtener_grupo_chat_id(request):
    """API para obtener el chat_id de un grupo desde las actualizaciones del bot"""
    if request.method == 'POST':
        try:
            api_base = get_api_base()
            if not api_base:
                return JsonResponse({'error': 'Bot Token no configurado'}, status=400)
            
            # Obtener las últimas actualizaciones
            updates = get_updates(api_base, limit=50)
            
            # Buscar grupos en las actualizaciones
            grupos_encontrados = []
            for update in reversed(updates):
                msg = update.get("message") or update.get("edited_message")
                if not msg:
                    continue
                
                chat = msg.get("chat")
                if chat and chat.get("type") == "group" and "id" in chat:
                    chat_id = chat["id"]
                    nombre_grupo = chat.get("title", "Sin nombre")
                    
                    # Evitar duplicados
                    if not any(g["chat_id"] == str(chat_id) for g in grupos_encontrados):
                        grupos_encontrados.append({
                            "chat_id": str(chat_id),
                            "nombre": nombre_grupo,
                            "username": chat.get("username", ""),
                        })
            
            if grupos_encontrados:
                return JsonResponse({
                    'success': True,
                    'grupos': grupos_encontrados,
                    'message': f'Se encontraron {len(grupos_encontrados)} grupo(s)'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'No se encontraron grupos. Asegúrate de que el bot esté en el grupo y que alguien haya enviado un mensaje recientemente.'
                })
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def user_upload_excel(request):
    """Subir usuarios desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        from sede.models import Sede
        from cargo.models import Cargo
        from dependencia.models import Dependencia
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Si username está vacío, saltar
                    continue
                
                username = str(row[0]).strip() if row[0] else None
                if not username:
                    continue
                
                email = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if email and email.lower() in ['none', 'null', '']:
                    email = None
                
                first_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if first_name and first_name.lower() in ['none', 'null', '']:
                    first_name = None
                
                last_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
                if last_name and last_name.lower() in ['none', 'null', '']:
                    last_name = None
                
                dni = str(row[4]).strip() if len(row) > 4 and row[4] else None
                if dni and dni.lower() in ['none', 'null', '']:
                    dni = None
                
                nombres = str(row[5]).strip() if len(row) > 5 and row[5] else None
                if nombres and nombres.lower() in ['none', 'null', '']:
                    nombres = None
                
                apellidos_paterno = str(row[6]).strip() if len(row) > 6 and row[6] else None
                if apellidos_paterno and apellidos_paterno.lower() in ['none', 'null', '']:
                    apellidos_paterno = None
                
                apellidos_materno = str(row[7]).strip() if len(row) > 7 and row[7] else None
                if apellidos_materno and apellidos_materno.lower() in ['none', 'null', '']:
                    apellidos_materno = None
                
                rol = str(row[8]).strip().lower() if len(row) > 8 and row[8] else 'usuario'
                if rol not in ['superadmin', 'admin', 'jefe_soporte', 'ingeniero_soporte', 'usuario']:
                    rol = 'usuario'
                
                sede_nombre = str(row[9]).strip() if len(row) > 9 and row[9] else None
                sede = None
                if sede_nombre:
                    try:
                        sede = Sede.objects.get(nombre=sede_nombre)
                    except Sede.DoesNotExist:
                        errors.append(f"Fila {row_idx}: Sede '{sede_nombre}' no encontrada")
                    except Sede.MultipleObjectsReturned:
                        sede = Sede.objects.filter(nombre=sede_nombre).first()
                
                cargo_nombre = str(row[10]).strip() if len(row) > 10 and row[10] else None
                cargo = None
                if cargo_nombre:
                    try:
                        cargo = Cargo.objects.get(nombre=cargo_nombre)
                    except Cargo.DoesNotExist:
                        errors.append(f"Fila {row_idx}: Cargo '{cargo_nombre}' no encontrado")
                    except Cargo.MultipleObjectsReturned:
                        cargo = Cargo.objects.filter(nombre=cargo_nombre).first()
                
                dependencia_nombre = str(row[11]).strip() if len(row) > 11 and row[11] else None
                dependencia = None
                if dependencia_nombre:
                    try:
                        dependencia = Dependencia.objects.get(nombre=dependencia_nombre)
                    except Dependencia.DoesNotExist:
                        errors.append(f"Fila {row_idx}: Dependencia '{dependencia_nombre}' no encontrada")
                    except Dependencia.MultipleObjectsReturned:
                        dependencia = Dependencia.objects.filter(nombre=dependencia_nombre).first()
                
                activo = True
                if len(row) > 12 and row[12] is not None:
                    activo_val = str(row[12]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            'email': email or f'{username}@example.com',
                            'first_name': first_name,
                            'last_name': last_name,
                            'dni': dni,
                            'nombres': nombres,
                            'apellidos_paterno': apellidos_paterno,
                            'apellidos_materno': apellidos_materno,
                            'rol': rol,
                            'sede': sede,
                            'cargo': cargo,
                            'dependencia': dependencia,
                            'activo': activo,
                            'is_active': activo,
                            'password': make_password('TempPassword123!')  # Password temporal
                        }
                    )
                    
                    if not created:
                        user.email = email or user.email
                        user.first_name = first_name or user.first_name
                        user.last_name = last_name or user.last_name
                        user.dni = dni or user.dni
                        user.nombres = nombres or user.nombres
                        user.apellidos_paterno = apellidos_paterno or user.apellidos_paterno
                        user.apellidos_materno = apellidos_materno or user.apellidos_materno
                        user.rol = rol
                        user.sede = sede
                        user.cargo = cargo
                        user.dependencia = dependencia
                        user.activo = activo
                        user.is_active = activo
                        user.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            df = pd.read_excel(excel_file)
            
            for idx, row in df.iterrows():
                username = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if not username:
                    continue
                
                email = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
                if email and email.lower() in ['none', 'null', '']:
                    email = None
                
                first_name = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None
                if first_name and first_name.lower() in ['none', 'null', '']:
                    first_name = None
                
                last_name = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else None
                if last_name and last_name.lower() in ['none', 'null', '']:
                    last_name = None
                
                dni = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else None
                if dni and dni.lower() in ['none', 'null', '']:
                    dni = None
                
                nombres = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else None
                if nombres and nombres.lower() in ['none', 'null', '']:
                    nombres = None
                
                apellidos_paterno = str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else None
                if apellidos_paterno and apellidos_paterno.lower() in ['none', 'null', '']:
                    apellidos_paterno = None
                
                apellidos_materno = str(row.iloc[7]).strip() if len(row) > 7 and pd.notna(row.iloc[7]) else None
                if apellidos_materno and apellidos_materno.lower() in ['none', 'null', '']:
                    apellidos_materno = None
                
                rol = str(row.iloc[8]).strip().lower() if len(row) > 8 and pd.notna(row.iloc[8]) else 'usuario'
                if rol not in ['superadmin', 'admin', 'jefe_soporte', 'ingeniero_soporte', 'usuario']:
                    rol = 'usuario'
                
                sede_nombre = str(row.iloc[9]).strip() if len(row) > 9 and pd.notna(row.iloc[9]) else None
                sede = None
                if sede_nombre:
                    try:
                        sede = Sede.objects.get(nombre=sede_nombre)
                    except Sede.DoesNotExist:
                        errors.append(f"Fila {idx + 2}: Sede '{sede_nombre}' no encontrada")
                    except Sede.MultipleObjectsReturned:
                        sede = Sede.objects.filter(nombre=sede_nombre).first()
                
                cargo_nombre = str(row.iloc[10]).strip() if len(row) > 10 and pd.notna(row.iloc[10]) else None
                cargo = None
                if cargo_nombre:
                    try:
                        cargo = Cargo.objects.get(nombre=cargo_nombre)
                    except Cargo.DoesNotExist:
                        errors.append(f"Fila {idx + 2}: Cargo '{cargo_nombre}' no encontrado")
                    except Cargo.MultipleObjectsReturned:
                        cargo = Cargo.objects.filter(nombre=cargo_nombre).first()
                
                dependencia_nombre = str(row.iloc[11]).strip() if len(row) > 11 and pd.notna(row.iloc[11]) else None
                dependencia = None
                if dependencia_nombre:
                    try:
                        dependencia = Dependencia.objects.get(nombre=dependencia_nombre)
                    except Dependencia.DoesNotExist:
                        errors.append(f"Fila {idx + 2}: Dependencia '{dependencia_nombre}' no encontrada")
                    except Dependencia.MultipleObjectsReturned:
                        dependencia = Dependencia.objects.filter(nombre=dependencia_nombre).first()
                
                activo = True
                if len(row) > 12 and pd.notna(row.iloc[12]):
                    activo_val = str(row.iloc[12]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            'email': email or f'{username}@example.com',
                            'first_name': first_name,
                            'last_name': last_name,
                            'dni': dni,
                            'nombres': nombres,
                            'apellidos_paterno': apellidos_paterno,
                            'apellidos_materno': apellidos_materno,
                            'rol': rol,
                            'sede': sede,
                            'cargo': cargo,
                            'dependencia': dependencia,
                            'activo': activo,
                            'is_active': activo,
                            'password': make_password('TempPassword123!')
                        }
                    )
                    
                    if not created:
                        user.email = email or user.email
                        user.first_name = first_name or user.first_name
                        user.last_name = last_name or user.last_name
                        user.dni = dni or user.dni
                        user.nombres = nombres or user.nombres
                        user.apellidos_paterno = apellidos_paterno or user.apellidos_paterno
                        user.apellidos_materno = apellidos_materno or user.apellidos_materno
                        user.rol = rol
                        user.sede = sede
                        user.cargo = cargo
                        user.dependencia = dependencia
                        user.activo = activo
                        user.is_active = activo
                        user.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def user_download_template(request):
    """Descargar template de Excel para usuarios"""
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Usuarios"
            
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            sheet['A1'] = 'Username'
            sheet['B1'] = 'Email'
            sheet['C1'] = 'First Name'
            sheet['D1'] = 'Last Name'
            sheet['E1'] = 'DNI'
            sheet['F1'] = 'Nombres'
            sheet['G1'] = 'Apellidos Paterno'
            sheet['H1'] = 'Apellidos Materno'
            sheet['I1'] = 'Rol'
            sheet['J1'] = 'Sede'
            sheet['K1'] = 'Cargo'
            sheet['L1'] = 'Dependencia'
            sheet['M1'] = 'Activo'
            
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            ejemplo_datos = [
                ['jperez', 'jperez@empresa.com', 'Juan', 'Pérez', '12345678', 'Juan', 'Pérez', 'García', 'usuario', 'Sede Central', 'Analista', 'Gerencia de TI', 1],
                ['mgarcia', 'mgarcia@empresa.com', 'María', 'García', '87654321', 'María', 'García', 'López', 'admin', 'Sede Lima', 'Gerente', 'Recursos Humanos', 1],
            ]
            
            for idx, (username, email, first_name, last_name, dni, nombres, apellidos_paterno, apellidos_materno, rol, sede, cargo, dependencia, activo) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = username
                sheet[f'B{idx}'] = email
                sheet[f'C{idx}'] = first_name
                sheet[f'D{idx}'] = last_name
                sheet[f'E{idx}'] = dni
                sheet[f'F{idx}'] = nombres
                sheet[f'G{idx}'] = apellidos_paterno
                sheet[f'H{idx}'] = apellidos_materno
                sheet[f'I{idx}'] = rol
                sheet[f'J{idx}'] = sede
                sheet[f'K{idx}'] = cargo
                sheet[f'L{idx}'] = dependencia
                sheet[f'M{idx}'] = activo
            
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 20
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 20
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 20
            sheet.column_dimensions['L'].width = 25
            sheet.column_dimensions['M'].width = 12
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_usuarios.xlsx"'
            workbook.save(response)
            return response
        else:
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Username': ['jperez', 'mgarcia'],
                'Email': ['jperez@empresa.com', 'mgarcia@empresa.com'],
                'First Name': ['Juan', 'María'],
                'Last Name': ['Pérez', 'García'],
                'DNI': ['12345678', '87654321'],
                'Nombres': ['Juan', 'María'],
                'Apellidos Paterno': ['Pérez', 'García'],
                'Apellidos Materno': ['García', 'López'],
                'Rol': ['usuario', 'admin'],
                'Sede': ['Sede Central', 'Sede Lima'],
                'Cargo': ['Analista', 'Gerente'],
                'Dependencia': ['Gerencia de TI', 'Recursos Humanos'],
                'Activo': [1, 1]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Usuarios')
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_usuarios.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )

@login_required
def profile(request):
    """Vista para editar el perfil del usuario"""
    user = request.user
    
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            # Asegurar que el usuario tenga username antes de guardar la foto
            if not user.username:
                user.save()  # Guardar primero para obtener username
            
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('accounts:profile')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ProfileForm(instance=user)
    
    return render(request, 'accounts/profile.html', {
        'form': form,
        'user': user
    })