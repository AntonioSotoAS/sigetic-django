from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from .models import Dependencia
from .forms import DependenciaForm
from sede.models import Sede

def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def dependencia_list(request):
    """Lista todas las dependencias"""
    dependencias = Dependencia.objects.all()
    return render(request, 'dependencia/dependencia_list.html', {'dependencias': dependencias})

@login_required
@user_passes_test(is_superuser)
def dependencia_create(request):
    """Crear nueva dependencia"""
    if request.method == 'POST':
        form = DependenciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dependencia creada exitosamente.')
            return redirect('dependencia:list')
    else:
        form = DependenciaForm()
    return render(request, 'dependencia/dependencia_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def dependencia_edit(request, pk):
    """Editar dependencia existente"""
    dependencia = get_object_or_404(Dependencia, pk=pk)
    if request.method == 'POST':
        form = DependenciaForm(request.POST, instance=dependencia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dependencia actualizada exitosamente.')
            return redirect('dependencia:list')
    else:
        form = DependenciaForm(instance=dependencia)
    return render(request, 'dependencia/dependencia_form.html', {'form': form, 'dependencia': dependencia, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def dependencia_delete(request, pk):
    """Eliminar dependencia"""
    dependencia = get_object_or_404(Dependencia, pk=pk)
    if request.method == 'POST':
        dependencia.delete()
        messages.success(request, 'Dependencia eliminada exitosamente.')
        return redirect('dependencia:list')
    return render(request, 'dependencia/dependencia_confirm_delete.html', {'dependencia': dependencia})

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def dependencia_upload_excel(request):
    """Subir dependencias desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        # Validar extensión
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        # Intentar importar openpyxl o pandas
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            # Usar openpyxl
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # Leer desde la fila 2 (asumiendo que la fila 1 tiene encabezados)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Si la primera columna está vacía, saltar
                    continue
                
                # Columna 1: Sede (opcional)
                sede_nombre = str(row[0]).strip() if len(row) > 0 and row[0] else None
                sede = None
                if sede_nombre:
                    try:
                        sede = Sede.objects.get(nombre=sede_nombre)
                    except Sede.DoesNotExist:
                        errors.append(f"Fila {row_idx}: Sede '{sede_nombre}' no encontrada")
                        continue
                    except Sede.MultipleObjectsReturned:
                        sede = Sede.objects.filter(nombre=sede_nombre).first()
                
                # Columna 2: Nombre (requerido)
                nombre = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not nombre:
                    errors.append(f"Fila {row_idx}: Nombre es requerido")
                    continue
                
                # Columna 3: Descripción (opcional)
                descripcion = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if descripcion and descripcion.lower() in ['none', 'null', '']:
                    descripcion = None
                
                # Columna 4: Activo (opcional, default True)
                activo = True
                if len(row) > 3 and row[3] is not None:
                    activo_val = str(row[3]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    dependencia, created = Dependencia.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'sede': sede,
                            'descripcion': descripcion,
                            'activo': activo
                        }
                    )
                    
                    if not created:
                        # Actualizar si ya existe
                        dependencia.sede = sede
                        dependencia.descripcion = descripcion
                        dependencia.activo = activo
                        dependencia.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            # Usar pandas
            df = pd.read_excel(excel_file)
            
            for idx, row in df.iterrows():
                # Columna 1: Sede (opcional)
                sede_nombre = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                sede = None
                if sede_nombre:
                    try:
                        sede = Sede.objects.get(nombre=sede_nombre)
                    except Sede.DoesNotExist:
                        errors.append(f"Fila {idx + 2}: Sede '{sede_nombre}' no encontrada")
                        continue
                    except Sede.MultipleObjectsReturned:
                        sede = Sede.objects.filter(nombre=sede_nombre).first()
                
                # Columna 2: Nombre (requerido)
                nombre = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
                if not nombre:
                    errors.append(f"Fila {idx + 2}: Nombre es requerido")
                    continue
                
                # Columna 3: Descripción (opcional)
                descripcion = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None
                if descripcion and descripcion.lower() in ['none', 'null', '']:
                    descripcion = None
                
                # Columna 4: Activo (opcional, default True)
                activo = True
                if len(row) > 3 and pd.notna(row.iloc[3]):
                    activo_val = str(row.iloc[3]).strip().lower()
                    activo = activo_val in ['1', 'true', 'activo', 'si', 'sí', 'yes']
                
                try:
                    dependencia, created = Dependencia.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'sede': sede,
                            'descripcion': descripcion,
                            'activo': activo
                        }
                    )
                    
                    if not created:
                        dependencia.sede = sede
                        dependencia.descripcion = descripcion
                        dependencia.activo = activo
                        dependencia.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]  # Limitar a 10 errores
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def dependencia_download_template(request):
    """Descargar template de Excel para dependencias"""
    try:
        # Intentar importar openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            # Crear workbook con openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Dependencias"
            
            # Estilos
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Encabezados
            sheet['A1'] = 'Sede'
            sheet['B1'] = 'Nombre'
            sheet['C1'] = 'Descripción'
            sheet['D1'] = 'Activo'
            
            # Aplicar estilos a los encabezados
            for cell in ['A1', 'B1', 'C1', 'D1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            # Datos de ejemplo
            ejemplo_datos = [
                ['Sede Central', 'Gerencia de TI', 'Departamento de Tecnología de la Información', 1],
                ['Sede Central', 'Recursos Humanos', 'Departamento de Recursos Humanos', 1],
                ['Sede Lima', 'Contabilidad', 'Departamento de Contabilidad', 1],
                ['Sede Lima', 'Administración', 'Departamento de Administración', 0],
            ]
            
            for idx, (sede, nombre, descripcion, activo) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = sede
                sheet[f'B{idx}'] = nombre
                sheet[f'C{idx}'] = descripcion
                sheet[f'D{idx}'] = activo
            
            # Ajustar ancho de columnas
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 50
            sheet.column_dimensions['D'].width = 12
            
            # Crear respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_dependencias.xlsx"'
            workbook.save(response)
            return response
        else:
            # Usar pandas
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Sede': ['Sede Central', 'Sede Central', 'Sede Lima', 'Sede Lima'],
                'Nombre': ['Gerencia de TI', 'Recursos Humanos', 'Contabilidad', 'Administración'],
                'Descripción': [
                    'Departamento de Tecnología de la Información',
                    'Departamento de Recursos Humanos',
                    'Departamento de Contabilidad',
                    'Departamento de Administración'
                ],
                'Activo': [1, 1, 1, 0]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Dependencias')
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_dependencias.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )