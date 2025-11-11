from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from .models import Sede
from .forms import SedeForm

def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def sede_list(request):
    """Lista todas las sedes"""
    sedes = Sede.objects.all()
    return render(request, 'sede/sede_list.html', {'sedes': sedes})

@login_required
@user_passes_test(is_superuser)
def sede_create(request):
    """Crear nueva sede"""
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sede creada exitosamente.')
            return redirect('sede:list')
    else:
        form = SedeForm()
    return render(request, 'sede/sede_form.html', {'form': form, 'action': 'Crear'})

@login_required
@user_passes_test(is_superuser)
def sede_edit(request, pk):
    """Editar sede existente"""
    sede = get_object_or_404(Sede, pk=pk)
    if request.method == 'POST':
        form = SedeForm(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sede actualizada exitosamente.')
            return redirect('sede:list')
    else:
        form = SedeForm(instance=sede)
    return render(request, 'sede/sede_form.html', {'form': form, 'sede': sede, 'action': 'Editar'})

@login_required
@user_passes_test(is_superuser)
def sede_delete(request, pk):
    """Eliminar sede"""
    sede = get_object_or_404(Sede, pk=pk)
    if request.method == 'POST':
        sede.delete()
        messages.success(request, 'Sede eliminada exitosamente.')
        return redirect('sede:list')
    return render(request, 'sede/sede_confirm_delete.html', {'sede': sede})

@login_required
@user_passes_test(is_superuser)
@require_http_methods(["POST"])
def sede_upload_excel(request):
    """Subir sedes desde archivo Excel"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No se encontró el archivo Excel'
            })
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
            })
        
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'message': 'Se requiere instalar openpyxl o pandas. Ejecuta: pip install openpyxl'
                })
        
        created_count = 0
        updated_count = 0
        errors = []
        
        if use_openpyxl:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                nombre = str(row[0]).strip() if row[0] else None
                if not nombre:
                    continue
                
                direccion = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if direccion and direccion.lower() in ['none', 'null', '']:
                    direccion = None
                
                telefono = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if telefono and telefono.lower() in ['none', 'null', '']:
                    telefono = None
                
                email = str(row[3]).strip() if len(row) > 3 and row[3] else None
                if email and email.lower() in ['none', 'null', '']:
                    email = None
                
                activa = True
                if len(row) > 4 and row[4] is not None:
                    activa_val = str(row[4]).strip().lower()
                    activa = activa_val in ['1', 'true', 'activa', 'si', 'sí', 'yes']
                
                try:
                    sede, created = Sede.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'direccion': direccion,
                            'telefono': telefono,
                            'email': email,
                            'activa': activa
                        }
                    )
                    
                    if not created:
                        sede.direccion = direccion
                        sede.telefono = telefono
                        sede.email = email
                        sede.activa = activa
                        sede.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {row_idx}: {str(e)}")
        else:
            df = pd.read_excel(excel_file)
            
            for idx, row in df.iterrows():
                nombre = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if not nombre:
                    continue
                
                direccion = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
                if direccion and direccion.lower() in ['none', 'null', '']:
                    direccion = None
                
                telefono = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None
                if telefono and telefono.lower() in ['none', 'null', '']:
                    telefono = None
                
                email = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else None
                if email and email.lower() in ['none', 'null', '']:
                    email = None
                
                activa = True
                if len(row) > 4 and pd.notna(row.iloc[4]):
                    activa_val = str(row.iloc[4]).strip().lower()
                    activa = activa_val in ['1', 'true', 'activa', 'si', 'sí', 'yes']
                
                try:
                    sede, created = Sede.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'direccion': direccion,
                            'telefono': telefono,
                            'email': email,
                            'activa': activa
                        }
                    )
                    
                    if not created:
                        sede.direccion = direccion
                        sede.telefono = telefono
                        sede.email = email
                        sede.activa = activa
                        sede.save()
                        updated_count += 1
                    else:
                        created_count += 1
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        message = f'Archivo procesado exitosamente.'
        if errors:
            message += f' Se encontraron {len(errors)} errores.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        })

@login_required
@user_passes_test(is_superuser)
def sede_download_template(request):
    """Descargar template de Excel para sedes"""
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                return HttpResponse(
                    'Se requiere instalar openpyxl. Ejecuta: pip install openpyxl',
                    content_type='text/plain',
                    status=500
                )
        else:
            use_openpyxl = True
        
        if use_openpyxl:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sedes"
            
            header_fill = PatternFill(start_color="c1121f", end_color="c1121f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            sheet['A1'] = 'Nombre'
            sheet['B1'] = 'Dirección'
            sheet['C1'] = 'Teléfono'
            sheet['D1'] = 'Email'
            sheet['E1'] = 'Activa'
            
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font
                sheet[cell].alignment = center_alignment
            
            ejemplo_datos = [
                ['Sede Central', 'Av. Principal 123', '987654321', 'central@empresa.com', 1],
                ['Sede Lima', 'Jr. Lima 456', '987654322', 'lima@empresa.com', 1],
                ['Sede Arequipa', 'Av. Arequipa 789', '987654323', 'arequipa@empresa.com', 1],
                ['Sede Trujillo', 'Av. Trujillo 321', '987654324', 'trujillo@empresa.com', 0],
            ]
            
            for idx, (nombre, direccion, telefono, email, activa) in enumerate(ejemplo_datos, start=2):
                sheet[f'A{idx}'] = nombre
                sheet[f'B{idx}'] = direccion
                sheet[f'C{idx}'] = telefono
                sheet[f'D{idx}'] = email
                sheet[f'E{idx}'] = activa
            
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 30
            sheet.column_dimensions['E'].width = 12
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_sedes.xlsx"'
            workbook.save(response)
            return response
        else:
            import pandas as pd
            from io import BytesIO
            
            datos = {
                'Nombre': ['Sede Central', 'Sede Lima', 'Sede Arequipa', 'Sede Trujillo'],
                'Dirección': ['Av. Principal 123', 'Jr. Lima 456', 'Av. Arequipa 789', 'Av. Trujillo 321'],
                'Teléfono': ['987654321', '987654322', '987654323', '987654324'],
                'Email': ['central@empresa.com', 'lima@empresa.com', 'arequipa@empresa.com', 'trujillo@empresa.com'],
                'Activa': [1, 1, 1, 0]
            }
            df = pd.DataFrame(datos)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sedes')
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_sedes.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(
            f'Error al generar template: {str(e)}',
            content_type='text/plain',
            status=500
        )
